import subprocess

# Seus códigos Python como strings
codigo1 = ""
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import io
import os
import openpyxl
import re


def process_pdfs(input_directory):
    """
    Processa todos os arquivos PDF em um diretório, extraindo texto e retornando um dicionário com o nome do arquivo e o texto extraído.
    """
    pdf_files = [f for f in os.listdir(input_directory) if f.lower().endswith('.pdf')]
    extracted_text = {}

    for filename in pdf_files:
        pdf_path = os.path.join(input_directory, filename)
        print(f"Processando {pdf_path}...")
        text = extract_text_from_pdf(pdf_path)
        extracted_text[filename] = text

    return extracted_text

def extract_text_from_pdf(pdf_path):
    """
    Extrai texto de um arquivo PDF usando PyMuPDF e Tesseract OCR para páginas com imagens.
    """
    text = ""
    try:
        pdf_document = fitz.open(pdf_path)
        for page_num in range(len(pdf_document)):
            page = pdf_document.load_page(page_num)
            page_text = page.get_text()
            if page_text.strip():
                text += page_text
            else:
                pix = page.get_pixmap()
                img = Image.open(io.BytesIO(pix.tobytes()))
                ocr_text = pytesseract.image_to_string(img, config='--psm x')
                text += ocr_text
        pdf_document.close()
        
        # Limpeza do texto
        text = text.replace('\n', ' ').replace('\r', ' ').strip()
        text = re.sub(' +', ' ', text)  # Remove múltiplos espaços
    except Exception as e:
        print(f"Erro ao processar o arquivo {pdf_path}: {e}")
    return text

def extract_data_from_text(text):
    """
    Extrai dados específicos do texto extraído do PDF usando expressões regulares.
    """
    data = {}

    patterns = {
        'cpf_cnpj': r'CPF/CNPJ[:\s]*([\d./-]+)',  # Ajustado para possível espaço após ':'.
        'razao_social': r'\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b',  # Nome da empresa.
        'uf': r'(?<=Belo Horizonte\s)([\w\s]+)',  # Ajustar se a UF estiver em uma posição específica.
        'municipio': r'\bMunicípio\s*:\s*(\w+)\b',  # Município antes de 'MG'.
        'endereco': r'(?<=RUA\s)([\w\s,]+)',  # Endereço começando com 'Rua'.
        'numero_documento': r'Número Documento:\s*(\d+)',
        'serie' : r'Série:\s*(\d+)',
        'data': r'Emitida em[:\s]*([\d/]+)',  # Data da emissão.
        'valor_dos_servicos': r'Valor dos serviços[:\s*R$ ]*([\d,.]+)',  # Valor dos serviços.
        'valor_descontos': r'Descontos[:\s*R$ ]*([\d,.]+)',  # Descontos.
        'valor_contabil': r'Valor Líquido[:\s*R$ ]*([\d,.]+)',  # Valor Líquido.
        'base_calculo': r'Base de Cálculo[:\s*R$ ]*([\d,.]+)',  # Base de Cálculo.
        'aliquota_iss': r'Alíquota[:\s]*([\d]+%)',  # Alíquota ISS.
        'valor_iss_normal': r'Valor do ISS[:\s*R$ ]*([\d,.]+)',  # Valor ISS Normal.
        'valor_iss_retido': r'ISS Retido na Fonte[:\s*R$ ]*([\d,.]+)',  # Valor ISS Retido.
        'valor_irrf': r'IR[:\s*R$ ]*([\d,.]+)',  # Valor IRRF.
        'valor_pis': r'PIS[:\s*R$ ]*([\d,.]+)',  # Valor PIS.
        'valor_cofins': r'COFINS[:\s*R$ ]*([\d,.]+)',  # Valor COFINS.
        'valor_csll': r'CSLL[:\s*R$ ]*([\d,.]+)',  # Valor CSLL.
        
        
    }

    for key, pattern in patterns.items():
        
        match = re.search(pattern, text, re.IGNORECASE)
        data[key] = match.group(1).strip() if match else "Não Encontrado"

    return data

def fill_excel_with_text_updated(text, template_excel_path, output_excel_path):
    wb = openpyxl.load_workbook(template_excel_path)
    sheet = wb.active

    for row, (filename, extracted_text) in enumerate(text.items(), start=3):
        data = extract_data_from_text(extracted_text)
        for col, key in enumerate([
            'cpf_cnpj', 'razao_social', 'uf', 'municipio', 'endereco','numero_documento', 'serie',            'data', 'valor_dos_servicos', 'valor_descontos', 'valor_contabil',
            'base_calculo', 'aliquota_iss', 'valor_iss_normal', 'valor_iss_retido',
            'valor_irrf', 'valor_pis', 'valor_cofins', 'valor_csll', 'valor_crf',
            'valor_inss'], start=1):  # Ajustar as colunas conforme o Excel
            sheet.cell(row=row, column=col, value=data.get(key, "Não Encontrado"))

    wb.save(output_excel_path)
    print(f"Excel preenchido salvo em {output_excel_path}")




def extract_cpf_cnpj(text):
    match = re.search(r'CPF/CNPJ:\s*([\d./-]+)', text)
    return match.group(1) if match else "Não Encontrado"

def extract_razao_social(text):
    """
    Extrai a Razão Social (nome da empresa) do texto extraído do PDF.
    O nome da empresa está localizado após o Código de Verificação e antes do CNPJ.
    """
    # Padrão para identificar o Código de Verificação
    codigo_verificacao_pattern = r'Código de Verificação:\s*([\w\d]+)'
    
    # Padrão para identificar o CNPJ
    cnpj_pattern = r'CPF/CNPJ:\s*\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}'
    
    # Encontrar o texto após o Código de Verificação e antes do CNPJ
    match_codigo_verificacao = re.search(codigo_verificacao_pattern, text)
    match_cnpj = re.search(cnpj_pattern, text)
    
    if match_codigo_verificacao and match_cnpj:
        start_index = match_codigo_verificacao.end()
        end_index = match_cnpj.start()
        
        # Captura o texto entre o Código de Verificação e o CNPJ
        text_between = text[start_index:end_index].strip()
        
        # Normaliza o texto para maiúsculas e remove possíveis quebras de linha
        empresa_name = text_between.replace('\n', ' ').strip().upper()
        
        return empresa_name if empresa_name else "RAZÃO SOCIAL NÃO ENCONTRADA."
    
    return "CNPJ NÃO ENCONTRADO NO TEXTO."

# Exemplo de uso
text = """
NFS-e - NOTA FISCAL DE SERVIÇOS ELETRÔNICA Nº:2024/9918 Emitida em: 15/07/2024 às 09:23:12 Competência: 15/07/2024 Código de Verificação: d18f199f PROTEGE PROTECAO E TRANSPORTE 
DE VALORES LTDA CPF/CNPJ: 43.035.146/0061-16 Inscrição Municipal: 0827308/002-X AVE PRESIDENTE CARLOS LUZ, 695, Caiçaras - Cep: 31230-000 Belo Horizonte MG Telefone:  (11)3156-0800 Email:   Tomador do(s) Serviço(s) CPF/CNPJ: 39.609.220/0001-52 Inscrição Municipal: Não Informado AGUIA V COMERCIO DE COMBUSTIVEIS LTDA AV V BARAO HOMEM DE MELO, 400, NOVA SUISSA - Cep: 30421-284 Belo Horizonte MG Telefone: Não Informado Email: Não Informado Discriminação do(s) Serviço(s) Servicos de processamento de numerario Nao incidencia de imposto na fonte conf. SC COSIT n 98 de 17/08/2018Vencimento da Fatura 20/08/2024 Valor aproximado de tributos:106.64 Código de Tributação do Município (CTISS) 1104-0/02-88 / Carga, descarga e arrumação de bens de qualquer espécie Subitem Lista de Serviços LC 116/03 / Descrição: 11.04 / Armazenamento, deposito, carga, descarga, arrumacao e guarda de bens de qualquer especie. Cod/Município da incidência do ISSQN: 3106200 / Belo Horizonte Natureza da Operação: Tributação no município Valor dos serviços: R$ 2.921,54 (-) Descontos: R$ 0,00 (-) Retenções Federais: R$ 0,00 (-) ISS Retido na Fonte: R$ 0,00 Valor Líquido: R$ 2.921,54 Valor dos serviços: R$ 2.921,54 (-) Deduções: R$ 0,00 (-) Desconto Incondicionado: R$ 0,00 (=) Base de Cálculo: R$ 2.921,54 (x) Alíquota: 5% (=)Valor do ISS: R$ 146,08 Retenções Federais: PIS: R$ 0,00 COFINS: R$ 0,00 IR: R$ 0,00 CSLL: R$ 0,00 INSS: R$ 0,00 Outras retenções: R$ 0,00 Outras Informações: Chave de acesso no Ambiente de Dados Nacional: 31062001243035146006116240000000991824077484851314. Prefeitura de Belo Horizonte - Secretaria Municipal de Fazenda Rua Espírito Santo, 605 - 3º andar - Centro - CEP: 30160-919 - Belo Horizonte MG. Dúvidas: SIGESP 02/08/2024, 08:16 :: NFS-e - Nota Fiscal de Serviços eletrônica :: https://bhissdigital.pbh.gov.br/nfse/pages/exibicaoNFS-e.jsf 1/1
"""

def extract_uf(text):
    pattern = r'(?<=Belo Horizonte\s)([\w\s]+)'
    match = re.search(pattern, text)
    print(f"Texto para busca: {text}")  # Mensagem de depuração
    print(f"Padrão de busca: {pattern}")  # Mensagem de depuração
    if match:
        print(f"Correspondência encontrada: {match.group(0)}")  # Mensagem de depuração
        return match.group(0).strip()
    else:
        print("Nenhuma correspondência encontrada.")  # Mensagem de depuração
        return "Não Encontrado"



def extract_municipio(text):
    """
    Extrai o município do texto extraído do PDF.
    O município está localizado antes da UF.
    """
    # Padrão para identificar a UF
    uf_pattern = r'(?<=\s)MG(?=\s|$)'  # Ajuste o padrão se a UF puder variar
    
    # Padrão para identificar o município
    municipio_pattern = r'Cod/Município da incidência do ISSQN:\s*(\d{7})\s*/\s*([^/]+)'

    # Encontrar a UF
    match_uf = re.search(uf_pattern, text)
    
    # Encontrar o município e o código do município
    match_municipio = re.search(municipio_pattern, text)
    
    if match_municipio:
        municipio_name = match_municipio.group(2).strip().upper()
        
        # Verifica se o município é encontrado antes da UF
        if match_uf:
            return municipio_name
        
    return "Município não encontrado."

# Exemplo de uso
text = """
NFS-e - NOTA FISCAL DE SERVIÇOS ELETRÔNICA Nº:2024/9918 Emitida em: 15/07/2024 às 09:23:12 Competência: 15/07/2024 Código de Verificação: d18f199f PROTEGE PROTECAO E TRANSPORTE 
DE VALORES LTDA CPF/CNPJ: 43.035.146/0061-16 Inscrição Municipal: 0827308/002-X AVE PRESIDENTE CARLOS LUZ, 695, Caiçaras - Cep: 31230-000 Belo Horizonte MG Telefone:  (11)3156-0800 Email:   Tomador do(s) Serviço(s) CPF/CNPJ: 39.609.220/0001-52 Inscrição Municipal: Não Informado AGUIA V COMERCIO DE COMBUSTIVEIS LTDA AV V BARAO HOMEM DE MELO, 400, NOVA SUISSA - Cep: 30421-284 Belo Horizonte MG Telefone: Não Informado Email: Não Informado Discriminação do(s) Serviço(s) Servicos de processamento de numerario Nao incidencia de imposto na fonte conf. SC COSIT n 98 de 17/08/2018Vencimento da Fatura 20/08/2024 Valor aproximado de tributos:106.64 Código de Tributação do Município (CTISS) 1104-0/02-88 / Carga, descarga e arrumação de bens de qualquer espécie Subitem Lista de Serviços LC 116/03 / Descrição: 11.04 / Armazenamento, deposito, carga, descarga, arrumacao e guarda de bens de qualquer especie. Cod/Município da incidência do ISSQN: 3106200 / Belo Horizonte Natureza da Operação: Tributação no município Valor dos serviços: R$ 2.921,54 (-) Descontos: R$ 0,00 (-) Retenções Federais: R$ 0,00 (-) ISS Retido na Fonte: R$ 0,00 Valor Líquido: R$ 2.921,54 Valor dos serviços: R$ 2.921,54 (-) Deduções: R$ 0,00 (-) Desconto Incondicionado: R$ 0,00 (=) Base de Cálculo: R$ 2.921,54 (x) Alíquota: 5% (=)Valor do ISS: R$ 146,08 Retenções Federais: PIS: R$ 0,00 COFINS: R$ 0,00 IR: R$ 0,00 CSLL: R$ 0,00 INSS: R$ 0,00 Outras retenções: R$ 0,00 Outras Informações: Chave de acesso no Ambiente de Dados Nacional: 31062001243035146006116240000000991824077484851314. Prefeitura de Belo Horizonte - Secretaria Municipal de Fazenda Rua Espírito Santo, 605 - 3º andar - Centro - CEP: 30160-919 - Belo Horizonte MG. Dúvidas: SIGESP 02/08/2024, 08:16 :: NFS-e - Nota Fiscal de Serviços eletrônica :: https://bhissdigital.pbh.gov.br/nfse/pages/exibicaoNFS-e.jsf 1/1
"""

print(extract_municipio(text))  # Esperado: "BELO HORIZONTE"


import re

def extract_endereco(text):
    """
    Extrai o endereço do texto extraído do PDF.
    O endereço está localizado entre a Inscrição Municipal e o Cep, excluindo o número da Inscrição Municipal e o número do CEP.
    """
    # Padrão para identificar a Inscrição Municipal e o Cep, capturando apenas o endereço
    pattern = r'Inscrição Municipal:\s*\d+[\w/-]*\s*(.*?)\s*-\s*Cep:\s*\d{5}-\d{3}'
    
    # Encontrar o endereço
    match = re.search(pattern, text, re.DOTALL)
    
    if match:
        endereco = match.group(1).strip()
        return endereco.upper()
    
    return "Endereço não encontrado."

# Exemplo de uso
text = """
NFS-e - NOTA FISCAL DE SERVIÇOS ELETRÔNICA Nº:2024/9918 Emitida em: 15/07/2024 às 09:23:12 Competência: 15/07/2024 Código de Verificação: d18f199f PROTEGE PROTECAO E TRANSPORTE DE VALORES LTDA CPF/CNPJ: 43.035.146/0061-16 Inscrição Municipal: 0827308/002-X AVE PRESIDENTE CARLOS LUZ, 695, Caiçaras - Cep: 31230-000 Belo Horizonte MG Telefone:  (11)3156-0800 Email:   Tomador do(s) Serviço(s) CPF/CNPJ: 39.609.220/0001-52 Inscrição Municipal: Não Informado AGUIA V COMERCIO DE COMBUSTIVEIS LTDA AV V BARAO HOMEM DE MELO, 400, NOVA SUISSA - Cep: 30421-284 Belo Horizonte MG Telefone: Não Informado Email: Não Informado Discriminação do(s) Serviço(s) Servicos de processamento de numerario Nao incidencia de imposto na fonte conf. SC COSIT n 98 de 17/08/2018Vencimento da Fatura 20/08/2024 Valor aproximado de tributos:106.64 Código de Tributação do Município (CTISS) 1104-0/02-88 / Carga, descarga e arrumação de bens de qualquer espécie Subitem Lista de Serviços LC 116/03 / Descrição: 11.04 / Armazenamento, deposito, carga, descarga, arrumacao e guarda de bens de qualquer especie. Cod/Município da incidência do ISSQN: 3106200 / Belo Horizonte Natureza da Operação: Tributação no município Valor dos serviços: R$ 2.921,54 (-) Descontos: R$ 0,00 (-) Retenções Federais: R$ 0,00 (-) ISS Retido na Fonte: R$ 0,00 Valor Líquido: R$ 2.921,54 Valor dos serviços: R$ 2.921,54 (-) Deduções: R$ 0,00 (-) Desconto Incondicionado: R$ 0,00 (=) Base de Cálculo: R$ 2.921,54 (x) Alíquota: 5% (=)Valor do ISS: R$ 146,08 Retenções Federais: PIS: R$ 0,00 COFINS: R$ 0,00 IR: R$ 0,00 CSLL: R$ 0,00 INSS: R$ 0,00 Outras retenções: R$ 0,00 Outras Informações: Chave de acesso no Ambiente de Dados Nacional: 31062001243035146006116240000000991824077484851314. Prefeitura de Belo Horizonte - Secretaria Municipal de Fazenda Rua Espírito Santo, 605 - 3º andar - Centro - CEP: 30160-919 - Belo Horizonte MG. Dúvidas: SIGESP 02/08/2024, 08:16 :: NFS-e - Nota Fiscal de Serviços eletrônica :: https://bhissdigital.pbh.gov.br/nfse/pages/exibicaoNFS-e.jsf 1/1
"""

print(extract_endereco(text))  # Esperado: "AVE PRESIDENTE CARLOS LUZ, 695, CAIÇARAS"

import re

def extract_numero_documento(text):
    """
    Extrai o número do documento do texto extraído do PDF.
    O número do documento está localizado antes da data de emissão e no formato Nº:XXXX/YYYY (com variáveis).
    """
    # Padrão para identificar o número do documento com partes de comprimento variável
    pattern = r'Nº:(\d+/[\d/]+)'
    
    # Encontrar o número do documento
    match = re.search(pattern, text)
    
    if match:
        numero_documento = match.group(1).strip()
        return numero_documento
    
    return "Número do documento não encontrado."

# Exemplo de uso
text = """
NFS-e - NOTA FISCAL DE SERVIÇOS ELETRÔNICA Nº:2024/9918 Emitida em: 15/07/2024 às 09:23:12 Competência: 15/07/2024 Código de Verificação: d18f199f PROTEGE PROTECAO E TRANSPORTE DE VALORES LTDA CPF/CNPJ: 43.035.146/0061-16 Inscrição Municipal: 0827308/002-X AVE PRESIDENTE CARLOS LUZ, 695, Caiçaras - Cep: 31230-000 Belo Horizonte MG Telefone:  (11)3156-0800 Email:   Tomador do(s) Serviço(s) CPF/CNPJ: 39.609.220/0001-52 Inscrição Municipal: Não Informado AGUIA V COMERCIO DE COMBUSTIVEIS LTDA AV V BARAO HOMEM DE MELO, 400, NOVA SUISSA - Cep: 30421-284 Belo Horizonte MG Telefone: Não Informado Email: Não Informado Discriminação do(s) Serviço(s) Servicos de processamento de numerario Nao incidencia de imposto na fonte conf. SC COSIT n 98 de 17/08/2018Vencimento da Fatura 20/08/2024 Valor aproximado de tributos:106.64 Código de Tributação do Município (CTISS) 1104-0/02-88 / Carga, descarga e arrumação de bens de qualquer espécie Subitem Lista de Serviços LC 116/03 / Descrição: 11.04 / Armazenamento, deposito, carga, descarga, arrumacao e guarda de bens de qualquer especie. Cod/Município da incidência do ISSQN: 3106200 / Belo Horizonte Natureza da Operação: Tributação no município Valor dos serviços: R$ 2.921,54 (-) Descontos: R$ 0,00 (-) Retenções Federais: R$ 0,00 (-) ISS Retido na Fonte: R$ 0,00 Valor Líquido: R$ 2.921,54 Valor dos serviços: R$ 2.921,54 (-) Deduções: R$ 0,00 (-) Desconto Incondicionado: R$ 0,00 (=) Base de Cálculo: R$ 2.921,54 (x) Alíquota: 5% (=)Valor do ISS: R$ 146,08 Retenções Federais: PIS: R$ 0,00 COFINS: R$ 0,00 IR: R$ 0,00 CSLL: R$ 0,00 INSS: R$ 0,00 Outras retenções: R$ 0,00 Outras Informações: Chave de acesso no Ambiente de Dados Nacional: 31062001243035146006116240000000991824077484851314. Prefeitura de Belo Horizonte - Secretaria Municipal de Fazenda Rua Espírito Santo, 605 - 3º andar - Centro - CEP: 30160-919 - Belo Horizonte MG. Dúvidas: SIGESP 02/08/2024, 08:16 :: NFS-e - Nota Fiscal de Serviços eletrônica :: https://bhissdigital.pbh.gov.br/nfse/pages/exibicaoNFS-e.jsf 1/1
"""

print(extract_numero_documento(text))  # Esperado: "2024/9918"


def extract_serie(text):
    pattern = r'Série:\s*(\d+)'
    match = re.search(pattern, text)
    return match.group(1).strip() if match else ""


def extract_data(text):
    match = re.search(r'Emitida em[:\s]*([\d/]+)', text)
    return match.group(1) if match else "Não Encontrado"

def extract_situacao(text):
    match = re.search(r'Situa[cç][aã]o:\s*(\d)', text)
    return match.group(1) if match else "0"

def extract_acumulador(text):
    match = re.search(r'Acumulador:\s*([\w\s]+)', text)
    result = match.group(1).strip() if match else ""
    print("Acumulador:", result)
    return result

def extract_cfop(text):
    match = re.search(r'CFOP:\s*(\d+)', text)
    result = match.group(1) if match else ""
    print("CFOP:", result)
    return result

def extract_valor_dos_servicos(text):
    match = re.search(r'Valor\s*dos\s*serviços[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else "Não Encontrado"
    print("Valor dos Serviços:", result)
    return result

def extract_valor_descontos(text):
    match = re.search(r'Descontos[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else "Não Encontrado"
    print("Valor Descontos:", result)
    return result

def extract_valor_contabil(text):
    match = re.search(r'Valor\s*Líquido[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else "Não Encontrado"
    print("Valor Contábil:", result)
    return result

def extract_base_calculo(text):
    match = re.search(r'Base\s*de\s*Cálculo[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else "Não Encontrado"
    print("Base de Calculo:", result)
    return result

def extract_aliquota_iss(text):
    match = re.search(r'Alíquota[:\s]*([\d]+%)', text)
    result = match.group(1) if match else ""
    print("Alíquota ISS:", result)
    return result

def extract_valor_iss_normal(text):
    match = re.search(r'Valor\s*do\s*ISS[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else ""
    print("Valor ISS Normal:", result)
    return result

def extract_valor_iss_retido(text):
    match = re.search(r'ISS\s*Retido\s*na\s*Fonte[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else ""
    print("Valor ISS Retido:", result)
    return result

def extract_valor_irrf(text):
    match = re.search(r'IR[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else ""
    print("Valor IRRF:", result)
    return result

def extract_valor_pis(text):
    match = re.search(r'PIS[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else ""
    print("Valor PIS:", result)
    return result

def extract_valor_cofins(text):
    match = re.search(r'COFINS[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else ""
    print("Valor COFINS:", result)
    return result

def extract_valor_csll(text):
    match = re.search(r'CSLL[:\s*R$ ]*([\d,.]+)', text)
    result = match.group(1) if match else ""
    print("Valor CSLL:", result)
    return result




def fill_excel_with_text_updated(text, template_excel_path, output_excel_path):
    wb = openpyxl.load_workbook(template_excel_path)
    sheet = wb.active

    for row, (filename, extracted_text) in enumerate(text.items(), start=3):
        sheet[f'A{row}'] = extract_cpf_cnpj(extracted_text)
        sheet[f'B{row}'] = extract_razao_social(extracted_text)
        sheet[f'C{row}'] = extract_uf(extracted_text)
        sheet[f'D{row}'] = extract_municipio(extracted_text)
        sheet[f'E{row}'] = extract_endereco(extracted_text)
        sheet[f'F{row}'] = extract_numero_documento(extracted_text)
        sheet[f'G{row}'] = extract_serie(extracted_text)
        sheet[f'H{row}'] = extract_data(extracted_text)
        sheet[f'I{row}'] = extract_situacao(extracted_text)
        sheet[f'J{row}'] = extract_acumulador(extracted_text)
        sheet[f'K{row}'] = extract_cfop(extracted_text)
        sheet[f'L{row}'] = extract_valor_dos_servicos(extracted_text)
        sheet[f'M{row}'] = extract_valor_descontos(extracted_text)
        sheet[f'N{row}'] = extract_valor_contabil(extracted_text)
        sheet[f'O{row}'] = extract_base_calculo(extracted_text)
        sheet[f'P{row}'] = extract_aliquota_iss(extracted_text)
        sheet[f'Q{row}'] = extract_valor_iss_normal(extracted_text)
        sheet[f'R{row}'] = extract_valor_iss_retido(extracted_text)
        sheet[f'S{row}'] = extract_valor_irrf(extracted_text)
        sheet[f'T{row}'] = extract_valor_pis(extracted_text)
        sheet[f'U{row}'] = extract_valor_cofins(extracted_text)
        sheet[f'V{row}'] = extract_valor_csll(extracted_text)
 

    wb.save(output_excel_path)
    print(f"Excel preenchido salvo em {output_excel_path}")


def main(input_directory, template_excel_path, output_excel_path):
    text_data = {}
    for filename in os.listdir(input_directory):
        if filename.endswith(".pdf"):
            pdf_path = os.path.join(input_directory, filename)
            print(f"Processando {pdf_path}...")
            text = extract_text_from_pdf(pdf_path)
            text_data[filename] = text

            # Exibe o texto extraído para análise
            print(f"Texto extraído de {filename}:")
            print(text)
            print("\n" + "="*50 + "\n")

    # Após análise, continue com o preenchimento do Excel
    fill_excel_with_text_updated(text_data, template_excel_path, output_excel_path)

if __name__ == "__main__":
    input_directory = 'c:\\Users\\jhennifer.nascimento\\nfs\\pdf\\st'
    template_excel_path = 'c:\\Users\\jhennifer.nascimento\\nfs\\modelo.xlsx.xlsx'
    output_excel_path = 'c:\\Users\\jhennifer.nascimento\\nfs\\output.xlsx'

    main(input_directory, template_excel_path, output_excel_path)

print("Executando código 1")


codigo2 = ""
import fitz  # PyMuPDF
import re
import pandas as pd

def extrair_texto_pdf(pdf_path):
    texto = ""
    with fitz.open(pdf_path) as pdf:
        for pagina in pdf:
            texto += pagina.get_text()
    return texto

def extrair_dados_nfse(texto):
    # Expressões regulares para capturar diferentes partes dos dados
    cnpj_pattern = r"CNPJ:\s*(\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2})"
    razao_social_pattern = r"Razão Social:\s*(.+)"
    uf_pattern = r"UF:\s*([A-Z]{2})"
    municipio_pattern = r"Município:\s*(.+)"
    endereco_pattern = r"Endereço:\s*(.+)"
    numero_doc_pattern = r"Número do Documento:\s*(\d+)"
    serie_pattern = r"Série:\s*(\d+)"
    data_pattern = r"Data:\s*(\d{2}/\d{2}/\d{4})"
    situacao_pattern = r"Situação:\s*(\d)"
    cfop_pattern = r"CFOP:\s*(\d+)"
    valores_pattern = r"Valor (Serviços|Descontos|Contábil|ISS|ISS Retido|IRRF|PIS|COFINS|CSLL):\s*R\$\s*([\d,.]+)"
    
    # Captura de CNPJ
    cnpj = re.search(cnpj_pattern, texto)
    cnpj = cnpj.group(1) if cnpj else None

    # Captura de Razão Social
    razao_social = re.search(razao_social_pattern, texto)
    razao_social = razao_social.group(1) if razao_social else None

    # Captura de UF
    uf = re.search(uf_pattern, texto)
    uf = uf.group(1) if uf else None

    # Captura de Município
    municipio = re.search(municipio_pattern, texto)
    municipio = municipio.group(1) if municipio else None

    # Captura de Endereço
    endereco = re.search(endereco_pattern, texto)
    endereco = endereco.group(1) if endereco else None

    # Captura de Número do Documento
    numero_doc = re.search(numero_doc_pattern, texto)
    numero_doc = numero_doc.group(1) if numero_doc else None

    # Captura de Série
    serie = re.search(serie_pattern, texto)
    serie = serie.group(1) if serie else None

    # Captura de Data
    data = re.search(data_pattern, texto)
    data = data.group(1) if data else None

    # Captura de Situação
    situacao = re.search(situacao_pattern, texto)
    situacao = situacao.group(1) if situacao else None

    # Captura de CFOP
    cfop = re.search(cfop_pattern, texto)
    cfop = cfop.group(1) if cfop else None

    # Captura de Valores Financeiros
    valores = {}
    for match in re.finditer(valores_pattern, texto):
        valores[match.group(1)] = match.group(2).replace('.', '').replace(',', '.')

    return {
        "CNPJ": cnpj,
        "Razão Social": razao_social,
        "Endereço": endereco,
        "Município": municipio,
        "UF": uf,
        "Número Documento": numero_doc,
        "Série": serie,
        "Data": data,
        "Situação": situacao,
        "CFOP": cfop,
        "Valores": valores
    }

def processar_pdf(pdf_path, excel_path):
    texto = extrair_texto_pdf(pdf_path)
    dados = extrair_dados_nfse(texto)
    
    # Transformando os valores financeiros em colunas separadas
    valores_df = pd.DataFrame([dados['Valores']])
    dados.pop('Valores')
    
    # Convertendo para DataFrame
    df_dados = pd.DataFrame([dados])
    
    # Combinando as informações em um único DataFrame
    df_final = pd.concat([df_dados, valores_df], axis=1)
    
    # Exportando para o Excel
    df_final.to_excel(excel_path, index=False, mode='a', header=False)

# Função principal para executar com subprocess
def main():
    pdf_path = 'C:\\Users\\jhennifer.nascimento\\nfs\\pdf\\st\\nfse_dominio.pdf'
    excel_path = 'C:\\Users\\jhennifer.nascimento\\nfs\\output.xlsx'
    
    # Executando o processamento de PDF e exportação para Excel em paralelo
    p1 = subprocess.Popen(['python', '-c', f'import sys; sys.path.append("."); from __main__ import processar_pdf; processar_pdf("{pdf_path}", "{excel_path}")'])
    
    # Se houver outro processo, adicione abaixo, por exemplo:
    # p2 = subprocess.Popen(['python', 'outro_script.py'])
    
    # Aguardando a conclusão do subprocess
    p1.wait()
    # p2.wait()  # Caso outro processo tenha sido iniciado

if __name__ == "__main__":
    main()

# Código 2
print("Executando código 2")



# Executar ambos os códigos em subprocessos
processo1 = subprocess.Popen(["python", "-c", codigo1], stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)
processo2 = subprocess.Popen(["python", "-c", codigo2], stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)

# Aguarda os dois processos terminarem e captura as saídas
stdout1, stderr1 = processo1.communicate()
stdout2, stderr2 = processo2.communicate()

# Mostra os resultados de ambos os processos
print("Saída do código 1:")
print(stdout1.decode())
print(stderr1.decode())

print("Saída do código 2:")
print(stdout2.decode())
print(stderr2.decode())
